#!/usr/bin/env python3
"""
МойСклад → Задолженность перед клиентами
Шаг 2: Формирование Excel-отчёта из pkl-кэша.

Запуск: python3 02_build_report.py
"""

import os, pickle
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Настройки ──────────────────────────────────────────────────────────────────
CACHE_PATH = os.environ.get('CACHE_PATH', '/tmp/moysklad_report_v4.pkl')
OUTPUT_PATH = os.environ.get('OUTPUT_PATH',
    os.path.expanduser('~/Nextcloud/0_Inbox/Задолженность_перед_клиентами_v2.xlsx'))

# ── Стили ──────────────────────────────────────────────────────────────────────
RUB = '#,##0.00" ₽"'
QTY = '#,##0'
THIN = Side(style='thin', color='D0D0D0')
THICK = Side(style='medium', color='2B4C7E')
TB = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

C = {
    'hdr': PatternFill(start_color='2B4C7E', end_color='2B4C7E', fill_type='solid'),
    'blue': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),
    'grey': PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid'),
    'green': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
    'orange': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'),
    'warn': PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'),
    'total': PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),
}
NONE_F = PatternFill(fill_type=None)

F = {
    'hdr': Font(name='Calibri', bold=True, color='FFFFFF', size=11),
    'bold': Font(name='Calibri', bold=True, size=11),
    'norm': Font(name='Calibri', size=10),
    'title_blue': Font(name='Calibri', bold=True, size=12, color='1565C0'),
    'title_warn': Font(name='Calibri', bold=True, size=12, color='E65100'),
    'title_dark': Font(name='Calibri', bold=True, size=12, color='2B4C7E'),
}

CAT_FILL = {
    'Бризер': C['green'],
    'Сплит/Кондиционер': C['orange'],
    'Прочее': C['grey'],
    'Услуга': NONE_F,
}

def set_cell(ws, row, col, value=None, font=None, fill=None, fmt=None,
             border=None, align=None, merge_to=None):
    c = ws.cell(row=row, column=col, value=value)
    if font: c.font = font
    if fill: c.fill = fill
    if fmt:  c.number_format = fmt
    if border: c.border = border
    if align: c.alignment = align
    if merge_to:
        ws.merge_cells(f'{get_column_letter(col)}{row}:{merge_to}{row}')
    return c

def hdr_row(ws, row, headers, fills=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F['hdr']
        c.fill = fills[i-1] if fills else C['hdr']
        c.border = TB
        c.alignment = Alignment(horizontal='center', wrap_text=True)

def freeze(ws, cell):
    ws.freeze_panes = cell

def autofilter(ws, row, cols):
    ws.auto_filter.ref = f'A{row}:{get_column_letter(cols)}{row}'

# ── Лист _Справочник ───────────────────────────────────────────────────────────
def build_spravochnik(wb, product_ref):
    ws = wb.create_sheet('_Справочник')
    ws.sheet_view.showGridLines = True
    hdr_row(ws, 1, ['Наименование', 'Категория', 'Производитель', 'Модель',
                     'buyPrice, ₽', 'kitPrice, ₽', 'Себест. за ед., ₽'])
    for col, w in enumerate([45, 18, 14, 18, 14, 14, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for i, (name, info) in enumerate(sorted(product_ref.items()), 2):
        fill = CAT_FILL.get(info.get('category', ''), NONE_F)
        for col, val in enumerate([
            name, info.get('category', ''), info.get('mfr', ''), info.get('model', ''),
            info.get('buy_price', 0), info.get('kit_price', 0),
        ], 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = fill; c.border = TB; c.font = F['norm']
            if col in [5, 6]: c.number_format = RUB
        # Formula for final cost = MAX(buyPrice, kitPrice)
        c = ws.cell(row=i, column=7, value=f'=MAX(E{i},F{i})')
        c.number_format = RUB; c.font = Font(name='Calibri', size=10, bold=True)
        c.fill = fill; c.border = TB
    freeze(ws, 'A2')
    autofilter(ws, 1, 7)
    return ws

# ── Лист _API_Клиенты ──────────────────────────────────────────────────────────
def build_api_clients(wb, clients):
    ws = wb.create_sheet('_API_Клиенты')
    hdr_row(ws, 1, ['Клиент', 'Код', 'Тел.', 'Тип', 'Баланс, ₽',
                     'Долг, ₽', 'Заказы', 'Кол-во заказов', 'Статус'])
    for col, w in enumerate([35, 12, 16, 12, 14, 14, 40, 12, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for i, (aid, info) in enumerate(clients.items(), 2):
        status = info.get('status', 'Закрытый')
        fill = NONE_F if status == 'Активный' else C['warn']
        vals = [
            info.get('name', ''),
            info.get('code', ''),
            info.get('phone', ''),
            'Юр. лицо' if info.get('companyType') == 'legal' else 'Физ. лицо',
            info.get('balance', 0),
            info.get('debt', 0),
            ', '.join(info.get('orders', [])),
            len(info.get('orders', [])),
            status,
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = fill; c.border = TB; c.font = F['norm']
            if col == 5: c.number_format = RUB
            if col == 6: c.number_format = RUB; c.font = Font(name='Calibri', size=10, bold=True)
    freeze(ws, 'A2')
    autofilter(ws, 1, 9)
    return ws

# ── Лист _API_Позиции ──────────────────────────────────────────────────────────
def build_api_positions(wb, results):
    ws = wb.create_sheet('_API_Позиции')
    hdr_row(ws, 1, ['Клиент', 'Код клиента', 'Заказ', 'Наименование товара',
                     'Кол-во', 'Долг аллоц., ₽', 'Категория', 'Статус'])
    for col, w in enumerate([30, 12, 16, 45, 8, 16, 18, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for i, r in enumerate(results, 2):
        fill = CAT_FILL.get(r.get('category', ''), NONE_F)
        if r.get('status') == 'Закрытый': fill = C['warn']
        vals = [r.get('client', ''), r.get('client_code', ''), r.get('order_name', ''),
                r.get('item_name', ''), r.get('qty', 0), r.get('debt_alloc', 0),
                r.get('category', ''), r.get('status', '')]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = fill; c.border = TB; c.font = F['norm']
            if col == 5: c.number_format = QTY; c.alignment = Alignment(horizontal='center')
            if col == 6: c.number_format = RUB
    freeze(ws, 'A2')
    autofilter(ws, 1, 8)
    return ws

# ── Лист Сводка ────────────────────────────────────────────────────────────────
def build_summary(wb, clients, results, generated_at):
    ws = wb.create_sheet('Сводка')
    n_active = sum(1 for v in clients.values() if v.get('status') == 'Активный')
    n_closed = sum(1 for v in clients.values() if v.get('status') == 'Закрытый')
    PLR = 1 + len(results)
    P = "'_API_Позиции'"
    Cl = "'_API_Клиенты'"
    S = "'_Справочник'"

    def cost_f(status, cat=None):
        cat_part = f'*({P}!G5:G{PLR}="{cat}")' if cat else ''
        return f'=SUMPRODUCT(({P}!H5:H{PLR}="{status}"){cat_part}*({P}!E5:E{PLR})*IFERROR(VLOOKUP({P}!D5:D{PLR},{S}!A:G,7,FALSE),0))'

    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1, value='Отчёт: Задолженность перед клиентами').font = Font(name='Calibri', bold=True, size=14, color='2B4C7E')
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')
    ws.cell(row=2, column=1, value=f'Дата: {generated_at[:10]}  |  Период: 01.01.2023 — {generated_at[:10]}  |  Долг = фактически оплачено − отгружено').font = F['norm']
    ws.merge_cells('A3:H3')
    ws.cell(row=3, column=1, value='Исключены: Микроклиматика, ИП Гончаров, Бризекс, тестовые контрагенты').font = F['norm']

    # Active clients section
    ws.merge_cells('A5:H5')
    ws.cell(row=5, column=1, value='АКТИВНЫЕ КЛИЕНТЫ (без аномалий)').font = F['title_blue']
    for i in range(1, 9): ws.cell(row=5, column=i).fill = C['blue']

    metrics = [
        (6, 'Клиентов с задолженностью', f'=COUNTIFS({Cl}!I:I,"Активный")', '0', True, None),
        (7, 'Общая задолженность', f'=SUMIFS({Cl}!F:F,{Cl}!I:I,"Активный")', RUB, True, None),
        (8, 'Общая себестоимость резерва', cost_f("Активный"), RUB, True, C['grey']),
        (9, 'Устройств в резерве (всего)', f'=SUMIFS({P}!E:E,{P}!H:H,"Активный")', QTY, True, None),
        (10, '  Бризеров', f'=SUMIFS({P}!E:E,{P}!H:H,"Активный",{P}!G:G,"Бризер")', QTY, False, C['grey']),
        (11, '  Сплит-систем / кондиционеров', f'=SUMIFS({P}!E:E,{P}!H:H,"Активный",{P}!G:G,"Сплит/Кондиционер")', QTY, False, None),
        (12, '  Прочих товаров', f'=SUMIFS({P}!E:E,{P}!H:H,"Активный",{P}!G:G,"Прочее")', QTY, False, C['grey']),
    ]
    for row, label, formula, fmt, bold, fill in metrics:
        ws.cell(row=row, column=1, value=label).font = Font(name='Calibri', size=11, bold=bold)
        c = ws.cell(row=row, column=3, value=formula); c.number_format = fmt; c.font = Font(name='Calibri', size=11, bold=bold)
        if fill:
            for j in range(1, 9): ws.cell(row=row, column=j).fill = fill

    # Category breakdown
    ws.merge_cells('A14:H14')
    ws.cell(row=14, column=1, value='РАЗБИВКА ПО КАТЕГОРИЯМ (активные клиенты)').font = F['title_dark']
    for i in range(1, 9): ws.cell(row=14, column=i).fill = C['blue']

    hdr_row(ws, 15, ['', 'Категория', 'Кол-во, шт', 'Долг, ₽', 'Себестоимость, ₽', 'Доля долга', '', ''])

    cats = [
        (16, 'Бризеры', 'Бризер', C['green']),
        (17, 'Сплит / Кондиционеры', 'Сплит/Кондиционер', C['orange']),
        (18, 'Прочие товары', 'Прочее', C['grey']),
    ]
    for row, label, cat, fill in cats:
        ws.cell(row=row, column=2, value=label).font = F['bold']; ws.cell(row=row, column=2).fill = fill; ws.cell(row=row, column=2).border = TB
        c = ws.cell(row=row, column=3, value=f'=SUMIFS({P}!E:E,{P}!H:H,"Активный",{P}!G:G,"{cat}")'); c.number_format = QTY; c.font = F['bold']; c.fill = fill; c.border = TB; c.alignment = Alignment(horizontal='center')
        c = ws.cell(row=row, column=4, value=f'=SUMIFS({P}!F:F,{P}!H:H,"Активный",{P}!G:G,"{cat}")'); c.number_format = RUB; c.font = F['bold']; c.fill = fill; c.border = TB
        c = ws.cell(row=row, column=5, value=cost_f("Активный", cat)); c.number_format = RUB; c.font = F['bold']; c.fill = fill; c.border = TB
        c = ws.cell(row=row, column=6, value=f'=IFERROR(D{row}/D21,0)'); c.number_format = '0.0%'; c.font = F['norm']; c.fill = fill; c.border = TB; c.alignment = Alignment(horizontal='center')

    # Услуги
    ws.cell(row=19, column=2, value='Услуги').font = Font(name='Calibri', size=11, italic=True); ws.cell(row=19, column=2).border = TB
    for col, formula, fmt in [(3, f'=SUMIFS({P}!E:E,{P}!H:H,"Активный",{P}!G:G,"Услуга")', QTY), (4, f'=SUMIFS({P}!F:F,{P}!H:H,"Активный",{P}!G:G,"Услуга")', RUB)]:
        c = ws.cell(row=19, column=col, value=formula); c.number_format = fmt; c.font = F['norm']; c.border = TB
        if col == 3: c.alignment = Alignment(horizontal='center')
    ws.cell(row=19, column=5, value='—').border = TB
    c = ws.cell(row=19, column=6, value='=IFERROR(D19/D21,0)'); c.number_format = '0.0%'; c.border = TB; c.alignment = Alignment(horizontal='center')

    # Total
    for j in range(2, 7):
        ws.cell(row=20, column=j).fill = C['total']
        ws.cell(row=20, column=j).font = F['bold']
        ws.cell(row=20, column=j).border = Border(top=THICK, bottom=THICK, left=THIN, right=THIN)
    ws.cell(row=20, column=2, value='ИТОГО')
    ws.cell(row=20, column=3, value='=SUM(C16:C19)').number_format = QTY
    ws.cell(row=20, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=20, column=4, value='=SUM(D16:D19)').number_format = RUB
    ws.cell(row=20, column=5, value='=SUM(E16:E18)').number_format = RUB

    # Anomalies
    ws.merge_cells('A22:H22')
    ws.cell(row=22, column=1, value='АНОМАЛИИ (закрытые заказы — НЕ входят в задолженность выше)').font = F['title_warn']
    for i in range(1, 9): ws.cell(row=22, column=i).fill = C['warn']

    anom = [
        (23, 'Клиентов', f'=COUNTIFS({Cl}!I:I,"Закрытый")', '0'),
        (24, 'Задолженность', f'=SUMIFS({Cl}!F:F,{Cl}!I:I,"Закрытый")', RUB),
        (25, 'Себестоимость', cost_f("Закрытый"), RUB),
        (26, 'Устройств', f'=SUMIFS({P}!E:E,{P}!H:H,"Закрытый")', QTY),
        (27, '  Бризеров', f'=SUMIFS({P}!E:E,{P}!H:H,"Закрытый",{P}!G:G,"Бризер")', QTY),
        (28, '  Сплит / Кондиционеров', f'=SUMIFS({P}!E:E,{P}!H:H,"Закрытый",{P}!G:G,"Сплит/Кондиционер")', QTY),
    ]
    for row, label, formula, fmt in anom:
        ws.cell(row=row, column=1, value=label).font = F['bold']
        c = ws.cell(row=row, column=3, value=formula); c.number_format = fmt; c.font = Font(name='Calibri', size=11, bold=True, color='E65100')

    # Top-10
    ws.merge_cells('A30:H30')
    ws.cell(row=30, column=1, value='ТОП-10 должников (активные)').font = F['title_dark']
    for i in range(1, 9): ws.cell(row=30, column=i).fill = C['blue']
    hdr_row(ws, 31, ['№', 'Клиент', 'Код', 'Тел.', 'Тип', 'Долг, ₽', 'Баланс, ₽', 'Заказов'])
    for idx in range(10):
        rr = 32 + idx; src = 5 + idx
        ws.cell(row=rr, column=1, value=idx+1).font = F['norm']; ws.cell(row=rr, column=1).border = TB; ws.cell(row=rr, column=1).alignment = Alignment(horizontal='center')
        for col, scol in [(2,'A'),(3,'B'),(4,'C'),(5,'D')]:
            c = ws.cell(row=rr, column=col, value=f'={Cl}!{scol}{src}'); c.font = F['norm']; c.border = TB
        c = ws.cell(row=rr, column=6, value=f'={Cl}!F{src}'); c.number_format = RUB; c.font = F['bold']; c.border = TB
        c = ws.cell(row=rr, column=7, value=f'={Cl}!E{src}'); c.number_format = RUB; c.font = F['norm']; c.border = TB
        c = ws.cell(row=rr, column=8, value=f'={Cl}!H{src}'); c.number_format = QTY; c.font = F['norm']; c.border = TB
        if idx % 2 == 1:
            for j in range(1, 9): ws.cell(row=rr, column=j).fill = C['grey']

    for col, w in [('A', 38), ('B', 36), ('C', 16), ('D', 18), ('E', 22), ('F', 14), ('G', 18), ('H', 12)]:
        ws.column_dimensions[col].width = w
    return ws

# ── Лист Детализация ───────────────────────────────────────────────────────────
def build_detail(wb, results, filter_status):
    name = 'Детализация' if filter_status == 'Активный' else 'Закрытые с долгом'
    ws = wb.create_sheet(name)
    hdr_row(ws, 1, ['Клиент', 'Код', 'Тел.', 'Заказ', 'Наименование товара',
                     'Категория', 'Кол-во', 'Долг аллоц., ₽', 'Себестоимость, ₽'])
    for col, w in enumerate([28, 12, 16, 14, 40, 18, 8, 16, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    S = "'_Справочник'"
    filtered = [r for r in results if r.get('status') == filter_status]
    for i, r in enumerate(filtered, 2):
        fill = CAT_FILL.get(r.get('category', ''), NONE_F)
        if filter_status == 'Закрытый': fill = C['warn']
        row_vals = [
            r.get('client', ''), r.get('client_code', ''), r.get('client_phone', ''),
            r.get('order_name', ''), r.get('item_name', ''), r.get('category', ''),
            r.get('qty', 0), r.get('debt_alloc', 0),
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = fill; c.border = TB; c.font = F['norm']
            if col == 7: c.number_format = QTY; c.alignment = Alignment(horizontal='center')
            if col == 8: c.number_format = RUB
        # Себестоимость = qty * VLOOKUP
        c = ws.cell(row=i, column=9)
        item_name = r.get('item_name', '').replace("'", "''")
        c.value = f'=G{i}*IFERROR(VLOOKUP(E{i},{S}!A:G,7,FALSE),0)'
        c.number_format = RUB; c.font = F['norm']; c.fill = fill; c.border = TB

    freeze(ws, 'A2')
    autofilter(ws, 1, 9)
    return ws

# ── Лист Бризеры ───────────────────────────────────────────────────────────────
def build_breezers(wb, results, product_ref):
    ws = wb.create_sheet('Бризеры')
    P = "'_API_Позиции'"
    S = "'_Справочник'"
    PLR = 1 + len(results)

    hdr = ['Производитель / Модель / Конфигурация', 'Кол-во, шт', 'Долг, ₽', 'Себестоимость, ₽']
    hdr_row(ws, 1, hdr)
    for col, w in enumerate([50, 12, 16, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    breezers = [r for r in results if r.get('category') == 'Бризер' and r.get('status') == 'Активный']
    by_mfr = defaultdict(lambda: defaultdict(list))
    for r in breezers:
        by_mfr[r.get('mfr') or 'Прочее'][r.get('model') or r.get('item_name', '')].append(r.get('item_name', ''))

    row = 2
    for mfr in sorted(by_mfr.keys(), key=lambda x: (x != 'AIRNANNY', x)):
        # Manufacturer total row
        mfr_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
        mfr_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
        ws.cell(row=row, column=1, value=f'  {mfr}').font = mfr_font
        for j in range(1, 5): ws.cell(row=row, column=j).fill = mfr_fill; ws.cell(row=row, column=j).border = TB
        c = ws.cell(row=row, column=2)
        c.value = f'=SUMPRODUCT(({P}!H5:H{PLR}="Активный")*({P}!G5:G{PLR}="Бризер")*ISNUMBER(SEARCH("{mfr}",{P}!D5:D{PLR}))*({P}!E5:E{PLR}))'
        c.number_format = QTY; c.font = mfr_font; c.alignment = Alignment(horizontal='center')
        c = ws.cell(row=row, column=3)
        c.value = f'=SUMPRODUCT(({P}!H5:H{PLR}="Активный")*({P}!G5:G{PLR}="Бризер")*ISNUMBER(SEARCH("{mfr}",{P}!D5:D{PLR}))*({P}!F5:F{PLR}))'
        c.number_format = RUB; c.font = mfr_font
        c = ws.cell(row=row, column=4)
        c.value = f'=SUMPRODUCT(({P}!H5:H{PLR}="Активный")*({P}!G5:G{PLR}="Бризер")*ISNUMBER(SEARCH("{mfr}",{P}!D5:D{PLR}))*({P}!E5:E{PLR})*IFERROR(VLOOKUP({P}!D5:D{PLR},{S}!A:G,7,FALSE),0))'
        c.number_format = RUB; c.font = mfr_font
        row += 1

        for model in sorted(by_mfr[mfr].keys()):
            # Model row
            model_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            ws.cell(row=row, column=1, value=f'    {model}').font = Font(name='Calibri', bold=True, size=11, color='1565C0')
            for j in range(1, 5): ws.cell(row=row, column=j).fill = model_fill; ws.cell(row=row, column=j).border = TB
            safe_model = model.replace('"', '""')
            c = ws.cell(row=row, column=2)
            c.value = f'=SUMPRODUCT(({P}!H5:H{PLR}="Активный")*({P}!G5:G{PLR}="Бризер")*ISNUMBER(SEARCH("{safe_model}",{P}!D5:D{PLR}))*({P}!E5:E{PLR}))'
            c.number_format = QTY; c.font = Font(name='Calibri', bold=True, size=11); c.alignment = Alignment(horizontal='center')
            c = ws.cell(row=row, column=3)
            c.value = f'=SUMPRODUCT(({P}!H5:H{PLR}="Активный")*({P}!G5:G{PLR}="Бризер")*ISNUMBER(SEARCH("{safe_model}",{P}!D5:D{PLR}))*({P}!F5:F{PLR}))'
            c.number_format = RUB; c.font = Font(name='Calibri', bold=True, size=11)
            c = ws.cell(row=row, column=4)
            c.value = f'=SUMPRODUCT(({P}!H5:H{PLR}="Активный")*({P}!G5:G{PLR}="Бризер")*ISNUMBER(SEARCH("{safe_model}",{P}!D5:D{PLR}))*({P}!E5:E{PLR})*IFERROR(VLOOKUP({P}!D5:D{PLR},{S}!A:G,7,FALSE),0))'
            c.number_format = RUB; c.font = Font(name='Calibri', bold=True, size=11)
            row += 1

            # Configurations
            unique_configs = sorted(set(by_mfr[mfr][model]))
            for cfg in unique_configs:
                safe_cfg = cfg.replace('"', '""')
                ws.cell(row=row, column=1, value=f'      {cfg}').font = F['norm']
                for j in range(1, 5): ws.cell(row=row, column=j).border = TB
                if row % 2 == 0:
                    for j in range(1, 5): ws.cell(row=row, column=j).fill = C['grey']
                c = ws.cell(row=row, column=2)
                c.value = f'=SUMPRODUCT(({P}!H5:H{PLR}="Активный")*({P}!D5:D{PLR}="{safe_cfg}")*({P}!E5:E{PLR}))'
                c.number_format = QTY; c.alignment = Alignment(horizontal='center')
                c = ws.cell(row=row, column=3)
                c.value = f'=SUMPRODUCT(({P}!H5:H{PLR}="Активный")*({P}!D5:D{PLR}="{safe_cfg}")*({P}!F5:F{PLR}))'
                c.number_format = RUB
                c = ws.cell(row=row, column=4)
                c.value = f'=SUMPRODUCT(({P}!H5:H{PLR}="Активный")*({P}!D5:D{PLR}="{safe_cfg}")*({P}!E5:E{PLR})*IFERROR(VLOOKUP({P}!D5:D{PLR},{S}!A:G,7,FALSE),0))'
                c.number_format = RUB
                row += 1

    freeze(ws, 'A2')
    return ws

# ── Main ───────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print(f'=== Building Excel Report ===')
    if not os.path.exists(CACHE_PATH):
        print(f'ERROR: Cache not found: {CACHE_PATH}')
        print('Run 01_fetch_data.py first')
        exit(1)

    with open(CACHE_PATH, 'rb') as f:
        data = pickle.load(f)

    clients = data['clients']
    results = data['results']
    product_ref = data['product_ref']
    generated_at = data.get('generated_at', datetime.now().isoformat())

    wb = Workbook()
    wb.remove(wb.active)

    build_summary(wb, clients, results, generated_at)
    build_spravochnik(wb, product_ref)
    build_api_clients(wb, clients)
    build_api_positions(wb, results)
    build_breezers(wb, results, product_ref)

    # Товары (все) — sheet with all items and filter
    ws_all = wb.create_sheet('Товары (все)')
    P = "'_API_Позиции'"
    S = "'_Справочник'"
    PLR = 1 + len(results)
    hdr_row(ws_all, 1, ['Наименование товара', 'Категория', 'Кол-во', 'Долг, ₽', 'Себестоимость, ₽'])
    for col, w in enumerate([45, 18, 10, 16, 18], 1):
        ws_all.column_dimensions[get_column_letter(col)].width = w
    unique_items = {}
    for r in results:
        if r.get('status') == 'Активный':
            name = r.get('item_name', '')
            cat = r.get('category', '')
            if name not in unique_items:
                unique_items[name] = cat
    for i, (name, cat) in enumerate(sorted(unique_items.items()), 2):
        safe = name.replace('"', '""')
        fill = CAT_FILL.get(cat, NONE_F)
        ws_all.cell(row=i, column=1, value=name).font = F['norm']; ws_all.cell(row=i, column=1).border = TB; ws_all.cell(row=i, column=1).fill = fill
        ws_all.cell(row=i, column=2, value=cat).font = F['norm']; ws_all.cell(row=i, column=2).border = TB; ws_all.cell(row=i, column=2).fill = fill
        c = ws_all.cell(row=i, column=3, value=f'=SUMPRODUCT(({P}!H5:H{PLR}="Активный")*({P}!D5:D{PLR}="{safe}")*({P}!E5:E{PLR}))'); c.number_format = QTY; c.font = F['norm']; c.border = TB; c.fill = fill; c.alignment = Alignment(horizontal='center')
        c = ws_all.cell(row=i, column=4, value=f'=SUMPRODUCT(({P}!H5:H{PLR}="Активный")*({P}!D5:D{PLR}="{safe}")*({P}!F5:F{PLR}))'); c.number_format = RUB; c.font = F['norm']; c.border = TB; c.fill = fill
        c = ws_all.cell(row=i, column=5, value=f'=C{i}*IFERROR(VLOOKUP(A{i},{S}!A:G,7,FALSE),0)'); c.number_format = RUB; c.font = F['norm']; c.border = TB; c.fill = fill
    freeze(ws_all, 'A2')
    ws_all.auto_filter.ref = f'A1:E1'

    build_detail(wb, results, 'Активный')
    build_detail(wb, results, 'Закрытый')

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)

    active_c = sum(1 for v in clients.values() if v.get('status') == 'Активный')
    closed_c = sum(1 for v in clients.values() if v.get('status') == 'Закрытый')
    active_r = [r for r in results if r.get('status') == 'Активный']
    closed_r = [r for r in results if r.get('status') == 'Закрытый']

    print(f'\n✅ SAVED: {OUTPUT_PATH}')
    print(f'   Sheets: {[s.title for s in wb.worksheets]}')
    print(f'   Клиентов: {active_c} активных + {closed_c} закрытых')
    print(f'   Позиций: {len(active_r)} активных + {len(closed_r)} закрытых')
    print(f'   Товаров в Справочнике: {len(product_ref)}')
